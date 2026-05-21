"""
Bake xlsx sources into Pipeline_Dashboard.html.

ALL DATA INPUT comes from OneDrive — this module never touches the host.
By the time bake() is invoked, the OneDrive downloader has populated
SOURCES_DIR (inside the container) with whatever it pulled.

Reads (paths inside the container):
  - {SOURCES_DIR}/Pipeline/Pipeline system team.xlsx   (sheet: "Pipeline";
                                                        falls back to "เทียบ")
  - {SOURCES_DIR}/SQ/YR*/SQ_*.xlsx                     (sheet: "SQ_Used_V103")

Writes (paths inside the container):
  {OUTPUT_PATH}      e.g. /app/static/Pipeline_Dashboard.html

The output HTML is built by taking the dashboard template baked into the
image at /app/template/Pipeline_Dashboard.template.html and substituting the
marker  __EMBEDDED_JSON__  with the freshly computed JSON.

This script is intentionally tolerant: any single xlsx that fails to parse
is logged and skipped, not fatal.
"""

from __future__ import annotations

import datetime as dt
import glob
import json
import logging
import os
import pathlib
import re
import sys
from typing import Any

import openpyxl

log = logging.getLogger("bake")

# ---------------------------------------------------------------------------
# Pipeline ("เทียบ" sheet)
# ---------------------------------------------------------------------------

_RESULT_MAP = {
    "win": "Win",
    "won": "Win",
    "lost": "Lost",
    "loss": "Lost",
    "drop": "Drop",
    "dropped": "Drop",
    "inactive": "Inactive",
    "in progress": "In Progress",
    "inprogress": "In Progress",
}


def _norm_result(v: Any) -> str:
    if v is None:
        return "In Progress"
    s = str(v).strip().lower()
    return _RESULT_MAP.get(s, "In Progress")


_Q_RE = re.compile(r"Q\s*([1-4])\s*/\s*(\d{4})", re.IGNORECASE)


def _parse_quarter(v: Any) -> tuple[int | None, int | None]:
    if v is None:
        return None, None
    s = str(v).strip()
    m = _Q_RE.search(s)
    if not m:
        return None, None
    return int(m.group(2)), int(m.group(1))


# --- Canonical brand normalization -----------------------------------------
# Maps the dozens of spelling, casing, and typo variants in the Pipeline
# xlsx to a single canonical brand name, so the dashboard's Top Products
# view rolls all variants up. Multi-brand strings (e.g. "Lenovo Huawei") are
# split into both canonical brands.
#
# Patterns are matched case-insensitively as substrings. Within each canonical
# brand, longer/more specific patterns should come first; across brands, the
# regex picks the LONGEST overall match at each position, so "HPE" wins over
# "HP" inside "HPE servers".
_BRAND_PATTERNS: list[tuple[str, tuple[str, ...]]] = [
    # Hardware vendors — product lines rolled up to parent brand.
    ("Dell",        ("dellemc", "dell emc", "dell hw", "dell server", "dell got",
                     "dell oem", "dell r740", "dell 2 server", "dell ุ6", "dell)",
                     "dell", "vxrail")),
    ("Huawei",      ("็huawei", "huawei", "hawei",  # 'hawei' = typo of Huawei
                     "dorado",                       # Huawei storage line
                     "hcs")),                        # Huawei Cloud Stack
    ("Lenovo",      ("lenovoo", "lenovo")),
    ("HPE",         ("hpe", "hewlett packard enterprise",
                     "dhci", "nimble", "synergy",    # HPE Nimble dHCI etc.
                     "proliant")),
    ("HP",          ("software (hp)", "hp")),
    ("Hitachi",     ("hitachi", "hds", "vsp")),
    ("IBM",         ("ibm",)),
    ("Brocade",     ("brocade", "braocade")),       # 'braocade' = typo
    ("Cisco",       ("cisco",)),
    ("Supermicro",  ("supermicro", "super micro")),
    ("ZTE",         ("zte",)),
    # Storage / virtualization software
    ("VMware",      ("wmware", "vmware", "vsan", "vvf")),
    ("NetApp",      ("netapp", "net app")),
    ("Pure Storage", ("purestroage", "purestorage", "pure storage", "pure")),
    ("Nutanix",     ("nutanix", "\bnkp\b")),
    # Backup / data management
    ("Veritas",     ("veritas", "netbackup", "backup exec",
                     "flex appliance", "infoscale")),
    ("Veeam",       ("veeam",)),
    ("Symantec",    ("symantec",)),
    # Networking / security
    ("Fortinet",    ("fortinet",)),
    ("Check Point", ("check point", "checkpoint", "check-point")),
    ("Palo Alto",   ("palo alto", "paloalto")),
    ("Juniper",     ("juniper",)),
    ("Arista",      ("arista",)),
    ("F5",          ("f5 ",)),
    ("Trend Micro", ("trend micro", "trendmicro")),
    # Enterprise software
    ("Microsoft",   ("microsoft", "office 365", "o365", "windows")),
    ("Oracle",      ("oracle", "oacle",            # 'oacle' = typo
                     "weblogic", "exadata", "exalogic")),
    ("Red Hat",     ("redhat", "red hat", "rhel")),
    ("Citrix",      ("citrix",)),
    ("Omnissa",     ("omnissa", "horizon vdi", "vdi")),
    # Local / regional vendors observed in Orders
    ("Atene",       ("atene",)),
    ("Vintcom",     ("vintcom",)),
    ("NeoConnect",  ("neoconnect", "neo connect")),
    ("Freewill",    ("freewill", "free will")),
    ("dcsstech",    ("dcsstech",)),
    ("Inet",        ("inet",)),
]


def _build_brand_regex():
    pair_to_canon: dict[str, str] = {}
    for canonical, patterns in _BRAND_PATTERNS:
        for p in patterns:
            pair_to_canon.setdefault(p.lower(), canonical)
    # Longest patterns first so the regex prefers "dell emc" over "dell",
    # "hpe" over "hp", etc.
    ordered = sorted(pair_to_canon.keys(), key=lambda s: -len(s))
    pattern = re.compile("|".join(re.escape(p) for p in ordered), re.IGNORECASE)
    return pattern, pair_to_canon


_BRAND_RE, _BRAND_CANONICAL = _build_brand_regex()


def _canonicalize_brands(raw: str) -> list[str]:
    """Find canonical brands present in raw text. Returns unique names in order."""
    if not raw:
        return []
    found: list[str] = []
    seen: set[str] = set()
    for m in _BRAND_RE.finditer(raw):
        canonical = _BRAND_CANONICAL[m.group(0).lower()]
        if canonical not in seen:
            seen.add(canonical)
            found.append(canonical)
    return found


def _first_canonical_brand(*texts: Any) -> str | None:
    """Return the first canonical brand mentioned in any of the given texts.

    Used to assign a single brand to a single SQ line item by looking at the
    item's remark and the order's project name. The texts are scanned in
    order; the earliest text with a brand wins.
    """
    for t in texts:
        if not t:
            continue
        names = _canonicalize_brands(str(t))
        if names:
            return names[0]
    return None


def _split_brands(product: Any) -> list[dict]:
    """Return brand attribution as a list of {name, share} for one product cell.

    Tries canonical normalization first (Dell EMC → Dell, Lenovoo → Lenovo, etc.).
    If no known brand is recognized, falls back to splitting the raw text on
    common delimiters so unknown vendors still appear in the dashboard.
    """
    if not product:
        return []
    raw = str(product).strip()
    canonicals = _canonicalize_brands(raw)
    if canonicals:
        share = round(1.0 / len(canonicals), 4)
        return [{"name": c, "share": share} for c in canonicals]
    # Unknown brand — keep existing splitter so the row still shows up under
    # whatever raw label the source used.
    parts = [p.strip() for p in re.split(r"[,/+&]| and | And | or | Or ", raw) if p.strip()]
    if not parts:
        parts = [raw]
    share = round(1.0 / len(parts), 4)
    return [{"name": p, "share": share} for p in parts]


def parse_pipeline(xlsx_path: pathlib.Path) -> list[dict]:
    log.info("pipeline: reading %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    # Prefer the "Pipeline" sheet (full historical data) over "เทียบ" (comparison view).
    sheet = "Pipeline" if "Pipeline" in wb.sheetnames else ("เทียบ" if "เทียบ" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Locate the header row dynamically — search the first 5 rows for a "Team" cell.
    header_idx = 0
    for i, r in enumerate(rows[:5]):
        if r and any(isinstance(v, str) and v.replace("\xa0", "").strip().lower() == "team" for v in r):
            header_idx = i
            break
    header = [str(h).strip() if h else "" for h in rows[header_idx]]
    rows = rows[header_idx + 1 :] + [None]  # sentinel so the loop's slice below works
    rows = rows[:-1]
    # Resolve column indices by header name (tolerant of NBSP/whitespace).
    def find(*aliases: str) -> int | None:
        for a in aliases:
            for i, h in enumerate(header):
                if h.replace("\xa0", "").strip().lower() == a.lower():
                    return i
        return None

    col_team = find("Team")
    col_sales = find("Sales")
    col_presales = find("Presales")
    col_customer = find("Customer")
    col_start = find("Start")
    col_target = find("Target")
    col_startdate = find("Start Date")
    col_project = find("Project name", "Project Name")
    col_projectcode = find("Project code", "Project Code")
    col_revenue = find("Revenue")
    col_product = find("Product")
    col_progress = find("%progress", "%Progress", "% progress")
    col_winpct = find("% win", "%win", "% Win")
    col_result = find("Win/Lost/Drop")
    col_status = find("Status")
    col_action = find("Action")

    out: list[dict] = []
    for r in rows:
        if not r or all(v in (None, "") for v in r):
            continue
        target = r[col_target] if col_target is not None else None
        start = r[col_start] if col_start is not None else None
        yr, qt = _parse_quarter(target)
        if yr is None:
            yr, qt = _parse_quarter(start)
        if yr is None:
            continue
        product = r[col_product] if col_product is not None else None
        revenue_raw = r[col_revenue] if col_revenue is not None else None
        try:
            revenue = float(revenue_raw) if revenue_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            revenue = 0.0
        winpct_raw = r[col_winpct] if col_winpct is not None else None
        try:
            winpct = float(winpct_raw) if winpct_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            winpct = 0.0
        out.append({
            "year": yr,
            "quarter": qt,
            "team": (str(r[col_team]).strip() if col_team is not None and r[col_team] else None),
            "result": _norm_result(r[col_result] if col_result is not None else None),
            "customer": (str(r[col_customer]).strip() if col_customer is not None and r[col_customer] else None),
            "sales": (str(r[col_sales]).strip() if col_sales is not None and r[col_sales] else None),
            "presales": (str(r[col_presales]).strip() if col_presales is not None and r[col_presales] else None),
            "product": (str(product).strip() if product else None),
            "productRaw": (str(product).strip() if product else None),
            "brands": _split_brands(product),
            "revenue": revenue,
            "project": (str(r[col_project]).strip() if col_project is not None and r[col_project] else None),
            "projectCode": (str(r[col_projectcode]).strip() if col_projectcode is not None and r[col_projectcode] else None),
            "status": (str(r[col_status]).strip() if col_status is not None and r[col_status] else None),
            "action": (str(r[col_action]).strip() if col_action is not None and r[col_action] else None),
            "winPct": winpct,
        })
    log.info("pipeline: %d rows parsed", len(out))
    return out


# ---------------------------------------------------------------------------
# SQ (Service Quotation) files
# ---------------------------------------------------------------------------

_SQ_NUM_RE = re.compile(r"(HI05000-\d{4}-\d{3})", re.IGNORECASE)


def _to_iso_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s or None


def _to_float(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_sq(xlsx_path: pathlib.Path, subfolder: str) -> dict | None:
    fname = xlsx_path.name
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        log.warning("sq: cannot open %s: %s", fname, exc)
        return None
    sheet = next((s for s in wb.sheetnames if s.startswith("SQ_Used")), wb.sheetnames[0])
    ws = wb[sheet]

    def cell(r: int, c: int) -> Any:
        return ws.cell(row=r, column=c).value

    # Header block (rows 5..11 in the standard SQ template).
    # Note: the "SQ No." cell (B5) is *not* trusted as authoritative — it's often
    # a stale template value copied from a previous year. The filename's
    # HI05000-YYYY-NNN segment is the source of truth (see below).
    sq_no_cell = cell(5, 7)
    customer = cell(6, 2)
    project_no = cell(6, 7)
    project_name = cell(7, 2)
    sale_team = cell(8, 2)
    sales_rep = cell(9, 2)
    type_of_sale = cell(9, 7)
    start_date = cell(10, 7)
    technology = cell(11, 2)
    end_date = cell(11, 7)

    # The FILENAME is the source of truth for both `sq_no` and `year`.
    # Order of trust (filename → subfolder → cell):
    #   1. Filename's HI05000-YYYY-NNN pattern — set when the file was created.
    #   2. The YR<YYYY> subfolder — only used if the filename has no match.
    #   3. The "SQ No." cell inside the spreadsheet — last resort, because
    #      it's often a stale template value copied from a previous year.
    sq_no = None
    year = None
    m = _SQ_NUM_RE.search(fname)
    if m:
        sq_no = m.group(1).upper()      # e.g. "HI05000-2025-044"
        year = int(re.search(r"-(\d{4})-", sq_no).group(1))
    if year is None:
        sub_m = re.fullmatch(r"YR(\d{4})", subfolder or "")
        if sub_m:
            year = int(sub_m.group(1))
    if sq_no is None and sq_no_cell:
        sq_no = str(sq_no_cell).strip() or None
    if year is None and sq_no_cell:
        cm = re.search(r"-(\d{4})-\d{3}", str(sq_no_cell))
        if cm:
            year = int(cm.group(1))

    # Walk rows looking for the three item sections.
    items: list[dict] = []
    section_label = None
    section_map = {
        "1.": "Services",
        "2.": "GP Product & External",
        "3.": "PR/PO Items",
    }
    # `sheet_total_*` = whatever the Grand Total row claims (kept for diagnostics).
    # The order's actual `total_*` fields are derived from the HI05000 item sum
    # below — keeps the dashboard KPIs consistent with the cost-center filter
    # the subtitle promises, and immune to stale/incorrect Grand Total cells.
    sheet_total_revenue = sheet_total_cost = sheet_total_gp = None

    max_row = ws.max_row
    for r in range(12, max_row + 1):
        a = cell(r, 1)
        c_col = cell(r, 3)
        sa = str(a).strip() if a else ""
        if sa:
            for prefix, lab in section_map.items():
                if sa.startswith(prefix):
                    section_label = lab
                    break
        if c_col and isinstance(c_col, str) and "Grand Total" in c_col:
            sheet_total_revenue = _to_float(cell(r, 4))
            sheet_total_cost = _to_float(cell(r, 5))
            sheet_total_gp = _to_float(cell(r, 6))
            continue
        if c_col and isinstance(c_col, str) and "Sub-Total" in c_col:
            continue
        # An item row has a value in column A that isn't a section header, and a
        # cost center value in column C.
        if not sa or sa.startswith(("1.", "2.", "3.", "4.")) or sa.lower().startswith("internal services") or sa.lower().startswith("product"):
            continue
        cost_center = c_col
        if not cost_center:
            continue
        revenue = _to_float(cell(r, 4))
        cost = _to_float(cell(r, 5))
        gp = _to_float(cell(r, 6))
        gp_pct = _to_float(cell(r, 7))
        tech = cell(r, 8)
        remark = cell(r, 11)
        remark_str = str(remark).strip() if remark else None
        # Canonical brand for the heatmap and product filter on the Orders tab.
        # Same _BRAND_PATTERNS the Pipeline page uses, so brand names are
        # consistent across both tabs. None means "no recognized brand" —
        # the JS detector will then fall back to item-code heuristics.
        brand = _first_canonical_brand(remark_str, project_name)
        items.append({
            "item": sa,
            "bu": str(cell(r, 2)).strip() if cell(r, 2) else None,
            "cost_center": str(cost_center).strip() if cost_center else None,
            "revenue": revenue,
            "cost": cost,
            "gp": gp,
            "gp_pct": gp_pct if gp_pct is not None else 0,
            "technology": str(tech).strip() if tech else None,
            "remark": remark_str,
            "ma_period": None,
            "section": section_label,
            "brand": brand,
        })

    # Normalize cost centers in place (uppercase, stripped) so dropdown
    # entries don't split on case alone. Typos like "HI0500" / "HI5000" are
    # preserved as-is so the user can spot them in the Cost Center dropdown.
    for it in items:
        cc = it.get("cost_center")
        if cc:
            it["cost_center"] = str(cc).upper().strip()

    # Keep ALL items in the parsed record. The Cost Center filter in the UI
    # (default HI05000) decides which items the heatmap/KPIs include.
    cost_centers = sorted({it["cost_center"] for it in items if it.get("cost_center")})

    # Compute the order's "default view" totals from HI05000 items only —
    # matches the dashboard's default Cost Center filter, and matches what
    # the subtitle promises. The JS recomputes from filtered items when the
    # user picks a different cost center.
    hi_items = [it for it in items if it.get("cost_center") and "HI05000" in it["cost_center"]]
    total_revenue = sum((it.get("revenue") or 0) for it in hi_items) or None
    total_cost = sum((it.get("cost") or 0) for it in hi_items) or None
    total_gp = sum((it.get("gp") or 0) for it in hi_items) or None
    # Diagnostic flag: True iff the Grand-Total cell disagrees with the sum
    # of HI05000 items by more than ฿1 (e.g. non-HI05000 items present, or a
    # stale Grand-Total formula like the OFFSA-style anomaly).
    sheet_mismatch = (
        sheet_total_revenue is not None
        and total_revenue is not None
        and abs(sheet_total_revenue - total_revenue) > 1
    )

    # Normalize Type of Sale: uppercase 2-3 letter abbreviations ("Si" → "SI"),
    # leave longer words as-is so "Sale" doesn't become "SALE".
    type_of_sale_str = str(type_of_sale).strip() if type_of_sale else None
    if type_of_sale_str and len(type_of_sale_str) <= 3 and type_of_sale_str.isalpha():
        type_of_sale_str = type_of_sale_str.upper()

    return {
        "file": fname,
        "year": year,
        "sq_no": str(sq_no).strip() if sq_no else None,
        "project_no": str(project_no).strip() if project_no else None,
        "project_name": str(project_name).strip() if project_name else None,
        "customer": str(customer).strip() if customer else None,
        "sale_team": str(sale_team).strip() if sale_team else None,
        "sales_rep": str(sales_rep).strip() if sales_rep else None,
        "type_of_sale": type_of_sale_str,
        "technology": str(technology).strip() if technology else None,
        "start_date": _to_iso_date(start_date),
        "end_date": _to_iso_date(end_date),
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "total_gp": total_gp,
        # The values written into the Grand Total row of the spreadsheet,
        # preserved for diagnostics. The dashboard KPIs do NOT use these.
        "sheet_total_revenue": sheet_total_revenue,
        "sheet_total_cost": sheet_total_cost,
        "sheet_total_gp": sheet_total_gp,
        "sheet_mismatch": sheet_mismatch,
        # Distinct cost centers present in this order's line items; the UI
        # uses the union across all orders to populate the Cost Center dropdown.
        "cost_centers": cost_centers,
        "items": items,
        "subfolder": subfolder,
    }


def collect_sq(sq_root: pathlib.Path) -> list[dict]:
    out: list[dict] = []
    for sub in sorted(sq_root.glob("YR*")):
        if not sub.is_dir():
            continue
        for xlsx in sorted(sub.glob("SQ_*.xlsx")):
            rec = parse_sq(xlsx, sub.name)
            if rec is not None:
                out.append(rec)
    log.info("sq: %d orders parsed", len(out))
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

EMBEDDED_MARKER = "__EMBEDDED_JSON__"


def build_embedded(sources: pathlib.Path) -> dict:
    pipeline_xlsx = sources / "Pipeline" / "Pipeline system team.xlsx"
    sq_root = sources / "SQ"
    pipeline = parse_pipeline(pipeline_xlsx) if pipeline_xlsx.exists() else []
    orders = collect_sq(sq_root) if sq_root.exists() else []
    return {
        "pipelineSource": pipeline_xlsx.name,
        "sqSources": [
            f"{o.get('subfolder','')}/{o.get('file','')}" for o in orders
        ],
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pipeline": pipeline,
        "orders": orders,
    }


def render_html(template_path: pathlib.Path, embedded: dict) -> str:
    tmpl = template_path.read_text(encoding="utf-8")
    if EMBEDDED_MARKER not in tmpl:
        raise RuntimeError(
            f"Template {template_path} is missing the {EMBEDDED_MARKER} marker. "
            "Run tools/extract_template.py to regenerate it from the original HTML."
        )
    json_text = json.dumps(embedded, ensure_ascii=False, separators=(", ", ": "))
    return tmpl.replace(EMBEDDED_MARKER, json_text)


def bake(
    sources: pathlib.Path,
    template: pathlib.Path,
    output: pathlib.Path,
    watermark: dict | None = None,
) -> dict:
    """Build the dashboard HTML and write a sibling last_bake.json.

    The bake_id in last_bake.json is monotonic — the dashboard JS uses it to
    detect "the server has re-baked since I loaded the page".
    """
    state_path = output.with_name("last_bake.json")
    prev_bake_id = 0
    if state_path.exists():
        try:
            prev_bake_id = int(json.loads(state_path.read_text()).get("bake_id", 0))
        except (ValueError, OSError, json.JSONDecodeError):
            prev_bake_id = 0
    bake_id = prev_bake_id + 1

    embedded = build_embedded(sources)
    embedded["bake_id"] = bake_id
    embedded["source_watermark"] = watermark or {}

    html = render_html(template, embedded)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(html, encoding="utf-8")

    state = {
        "bake_id": bake_id,
        "generated_at": embedded["generated_at"],
        "pipeline_rows": len(embedded["pipeline"]),
        "orders": len(embedded["orders"]),
        "source_watermark": watermark or {},
        "output": str(output),
    }
    state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")

    return state


def _main(argv: list[str]) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    sources = pathlib.Path(os.environ.get("SOURCES_DIR", "/app/sources"))
    template = pathlib.Path(os.environ.get("TEMPLATE_PATH", "/app/template/Pipeline_Dashboard.template.html"))
    output = pathlib.Path(os.environ.get("OUTPUT_PATH", "/app/static/Pipeline_Dashboard.html"))
    summary = bake(sources, template, output)
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv))
